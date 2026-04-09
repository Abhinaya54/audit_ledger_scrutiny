# scrutiny/exporter.py
# Excel export helper with layout formatting for readability.
# Preserves incoming DataFrame column names and order as provided.

import io
import re
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


CATEGORY_COL = "Scrutiny Category"
REASON_COL = "Scrutiny Reason"


def _apply_layout(ws, frame: pd.DataFrame) -> None:
    """Apply consistent styling and column sizing for a readable Excel layout."""
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    # Use a bounded sample while sizing to keep export fast on large workbooks.
    sampled = frame.head(5000)
    for idx, col_name in enumerate(frame.columns, start=1):
        col_letter = get_column_letter(idx)
        max_len = len(str(col_name))

        for value in sampled.iloc[:, idx - 1]:
            if pd.isna(value):
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _category_summary(detections: pd.DataFrame) -> pd.DataFrame:
    if detections.empty:
        return pd.DataFrame(columns=["Category", "Count"])

    counts = (
        detections[CATEGORY_COL]
        .fillna("")
        .astype(str)
        .str.split(",")
        .explode()
        .str.strip()
    )
    counts = counts[counts != ""]

    if counts.empty:
        return pd.DataFrame(columns=["Category", "Count"])

    summary = counts.value_counts().reset_index()
    summary.columns = ["Category", "Count"]
    return summary


def _sheet_name_for_category(category: str, used: set[str]) -> str:
    overrides = {
        "Weak Narration": "Weak_Narrations",
        "Round Numbers": "Round_Amounts",
        "Duplicate Check": "Duplicates",
        "ML Anomaly": "High_Risk",
        "Manual Journal": "Manual_Journal",
        "Weekend Entries": "Weekend_Entries",
        "Period End": "Period_End",
    }

    base = overrides.get(category, category)
    base = re.sub(r"[^A-Za-z0-9]+", "_", str(base)).strip("_") or "Category"
    base = base[:31]

    if base not in used:
        used.add(base)
        return base

    idx = 2
    while True:
        suffix = f"_{idx}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def _filter_category_rows(detections: pd.DataFrame, category: str) -> pd.DataFrame:
    category_tokens = (
        detections[CATEGORY_COL]
        .fillna("")
        .astype(str)
        .str.split(",")
        .apply(lambda tokens: [t.strip() for t in tokens if t.strip()])
    )
    mask = category_tokens.apply(lambda tokens: category in tokens)
    return detections[mask].copy()


def export(df: pd.DataFrame, output_path: str = None) -> bytes:
    """Generate a multi-sheet scrutiny workbook with readable formatting."""

    out = df.copy().reset_index(drop=True)
    if CATEGORY_COL not in out.columns:
        out[CATEGORY_COL] = ""
    if REASON_COL not in out.columns:
        out[REASON_COL] = ""

    detections = out[out[CATEGORY_COL].fillna("").astype(str).str.strip() != ""].copy()
    summary = _category_summary(detections)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Original_Data")
        _apply_layout(writer.sheets["Original_Data"], out)

        detections.to_excel(writer, index=False, sheet_name="All_Detections")
        _apply_layout(writer.sheets["All_Detections"], detections)

        used_sheet_names: set[str] = {"Original_Data", "All_Detections", "Summary"}
        for category in summary["Category"].tolist():
            category_df = _filter_category_rows(detections, category)
            sheet_name = _sheet_name_for_category(category, used_sheet_names)
            category_df.to_excel(writer, index=False, sheet_name=sheet_name)
            _apply_layout(writer.sheets[sheet_name], category_df)

        summary.to_excel(writer, index=False, sheet_name="Summary")
        _apply_layout(writer.sheets["Summary"], summary)

    buf.seek(0)
    result = buf.read()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(result)

    return result
