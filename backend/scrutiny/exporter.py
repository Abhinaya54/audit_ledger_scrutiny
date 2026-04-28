# scrutiny/exporter.py
# Excel export helper with layout formatting for readability.
# Preserves incoming DataFrame column names and order as provided.

import io
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


CATEGORY_COL = "Anomaly_Type"
REASON_COL = "Reason"


def _apply_layout(ws, frame: pd.DataFrame) -> None:
    """Apply consistent styling and column sizing for a readable Excel layout."""
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    # Use a bounded sample while sizing to keep export fast on large workbooks.
    sampled = frame.head(5000)
    for idx, col_name in enumerate(frame.columns, start=1):
        col_letter = get_column_letter(idx)
        max_len = len(str(col_name))

        for value in sampled.iloc[:, idx - 1]:
            if pd.isna(value):
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _category_summary(detections: pd.DataFrame) -> pd.DataFrame:
    if detections.empty:
        return pd.DataFrame(columns=["Category", "Count"])

    counts = (
        detections[CATEGORY_COL]
        .fillna("")
        .astype(str)
        .str.split(",")
        .explode()
        .str.strip()
    )
    counts = counts[counts != ""]

    if counts.empty:
        return pd.DataFrame(columns=["Category", "Count"])

    summary = counts.value_counts().reset_index()
    summary.columns = ["Anomaly_Type", "Count"]
    return summary


def export(df: pd.DataFrame, output_path: str = None) -> bytes:
    """Generate a two-sheet audit workbook with suspicious transactions and summary."""

    out = df.copy().reset_index(drop=True)
    if CATEGORY_COL not in out.columns:
        out[CATEGORY_COL] = ""
    if REASON_COL not in out.columns:
        out[REASON_COL] = ""

    detections = out[out[CATEGORY_COL].fillna("").astype(str).str.strip() != ""].copy()
    summary = _category_summary(detections)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        detections.to_excel(writer, index=False, sheet_name="Suspicious_Transactions")
        _apply_layout(writer.sheets["Suspicious_Transactions"], detections)

        summary.to_excel(writer, index=False, sheet_name="Summary")
        _apply_layout(writer.sheets["Summary"], summary)

    buf.seek(0)
    result = buf.read()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(result)

    return result
