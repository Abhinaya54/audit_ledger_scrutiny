# scrutiny/exporter.py
# Excel report exporter.
# Takes the flagged DataFrame and produces a styled .xlsx file with:
# - Mandatory disclaimer at the top
# - All original GL columns preserved
# - Scrutiny Category and Scrutiny Reason as the final two columns
# - Styled: bold header, alternating row colours, auto-fit columns

import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

DISCLAIMER = (
    "Inclusion of a transaction in the Ledger Scrutiny Register does not indicate an error, "
    "fraud, or misstatement. It indicates that the transaction meets predefined scrutiny "
    "criteria and requires auditor review."
)

# Colours
HEADER_FILL    = PatternFill("solid", fgColor="1A376C")
DISCLAIMER_FILL= PatternFill("solid", fgColor="FFF3CD")
ROW_FILL_ODD   = PatternFill("solid", fgColor="F2F6FC")
ROW_FILL_EVEN  = PatternFill("solid", fgColor="FFFFFF")
CATEGORY_FILL  = PatternFill("solid", fgColor="E8F4FD")

WHITE_FONT     = Font(color="FFFFFF", bold=True, size=10)
DISCLAIMER_FONT= Font(color="7D5700", bold=True, italic=True, size=9)
BODY_FONT      = Font(size=9)
CATEGORY_FONT  = Font(color="1A376C", bold=True, size=9)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export(df: pd.DataFrame, output_path: str = None) -> bytes:
    """
    Export the flagged GL rows to a styled Excel file.

    Args:
        df          : DataFrame containing ALL rows (flagged and unflagged).
                      Only rows where scrutiny_flag == True are exported.
        output_path : if provided, saves the file to this path.
                      Always returns the file as bytes for Streamlit download.

    Returns:
        bytes — the .xlsx file content
    """
    # ── Filter: keep only flagged rows ────────────────────────────────────────
    flagged = df[df["scrutiny_flag"]].copy().reset_index(drop=True)

    if flagged.empty:
        # Return an empty sheet with a message
        flagged = pd.DataFrame({"Message": ["No transactions flagged for scrutiny."]})

    # ── Reorder columns: original first, then scrutiny columns ───────────────
    scrutiny_cols = ["scrutiny_category", "scrutiny_reason"]
    ml_cols       = [c for c in ["ml_anomaly_flag", "ml_anomaly_score"] if c in flagged.columns]
    other_cols    = [c for c in flagged.columns
                     if c not in scrutiny_cols + ml_cols + ["scrutiny_flag"]]
    final_cols    = other_cols + scrutiny_cols + ml_cols
    final_cols    = [c for c in final_cols if c in flagged.columns]
    flagged       = flagged[final_cols]

    # ── Write to bytes buffer ─────────────────────────────────────────────────
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        flagged.to_excel(writer, sheet_name="Scrutiny Register", index=False, startrow=2)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Scrutiny Register"]

    # ── Disclaimer row (row 1) ────────────────────────────────────────────────
    ws.insert_rows(1)
    ws.insert_rows(1)
    disclaimer_cell = ws.cell(row=1, column=1)
    disclaimer_cell.value = disclaimer_cell.value = DISCLAIMER
    disclaimer_cell.font  = DISCLAIMER_FONT
    disclaimer_cell.fill  = DISCLAIMER_FILL
    disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1,   end_column=len(final_cols)
    )
    ws.row_dimensions[1].height = 40

    # ── Header row (row 3 after insertions) ───────────────────────────────────
    header_row = 3
    for col_idx, col_name in enumerate(final_cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx in range(header_row + 1, ws.max_row + 1):
        fill = ROW_FILL_ODD if (row_idx % 2 == 0) else ROW_FILL_EVEN
        for col_idx in range(1, len(final_cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Highlight scrutiny category and reason columns in light blue
            col_name = final_cols[col_idx - 1]
            if col_name in ("scrutiny_category", "scrutiny_reason"):
                cell.fill = CATEGORY_FILL
                cell.font = CATEGORY_FONT

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col_idx, col_name in enumerate(final_cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(col_name)),
            *[
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(header_row, min(ws.max_row + 1, header_row + 50))
            ],
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Save ──────────────────────────────────────────────────────────────────
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    excel_bytes = output_buffer.read()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(excel_bytes)
        print(f"Report saved: {output_path}  ({len(flagged):,} flagged rows)")

    return excel_bytes
